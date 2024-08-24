import turtle
import random
import openpyxl
from english_words import english_words_set
import requests
import tkinter



str1 = """Guess the Wordle in 6 tries:
* Each guess must be a valid 4/5/6-letter word.
* After a guess, the tiles will change color.
* The color of the tiles will change to show how close your guess was to the word.
* If the tile turns gray then the letter is not present the word.
* If the tile turn yellow then the letter is present in the word but not in the correct position.
* If the tile turns green then the letter is in the correct position.
"""
turtle.setup(width=0.999,height=0.999)
food = 0
movies = 0
places = 0
sports = 0
game = 0
mode = 0
mode1 = 0
wb = openpyxl.load_workbook(r'excel/text.xlsx')
ws = wb['Sheet1']

list1 = list(set(list(ws.iter_cols(max_col=1, values_only=True))[0]))
list1.remove(None)
list2 = list(set(list(ws.iter_cols(max_col=7, values_only=True))[2]))
list2.remove(None)
list3 = list(set(list(ws.iter_cols(max_col=5, values_only=True))[4]))
list3.remove(None)
list4 = list(set(list(ws.iter_cols(max_col=7, values_only=True))[6]))
list4.remove(None)
list5 = list(set(list(ws.iter_cols(max_col=9, values_only=True))[7]))
list5.remove(None)
list6 = list(set(list(ws.iter_cols(max_col=11, values_only=True))[8]))
list6.remove(None)

list7 = list(set(list(ws.iter_cols(max_col=13, values_only=True))[10]))
list7.remove(None)

list8 = list(set(list(ws.iter_cols(max_col=13, values_only=True))[11]))
list8.remove(None)
list9 = list(set(list(ws.iter_cols(max_col=13, values_only=True))[12]))
list9.remove(None)
list10 = list(set(list(ws.iter_cols(max_col=17, values_only=True))[14]))
list10.remove(None)
list11 = list(set(list(ws.iter_cols(max_col=18, values_only=True))[15]))
list11.remove(None)
list12 = list(set(list(ws.iter_cols(max_col=18, values_only=True))[16]))
list12.remove(None)
list13 = list(set(list(ws.iter_cols(max_col=19, values_only=True))[18]))
list13.remove(None)
list14 = list(set(list(ws.iter_cols(max_col=20, values_only=True))[19]))
list14.remove(None)
list15 = list(set(list(ws.iter_cols(max_col=21, values_only=True))[20]))
list15.remove(None)

random_s = ''
a2=-200
p3=-190
p4=40
size2=40
ttl = turtle.Turtle()
ttl2 = turtle.Turtle()
ttl3 = turtle.Turtle()
ttl1 = turtle.Turtle()
ttl4 = turtle.Turtle()
def gamemode():
    global mode
    tr = turtle.Turtle()
    tr.penup()
    tr.screen.addshape(r'excel/kn.gif')
    tr.goto(-200,355)
    tr.pendown()
    tr.shape(r'excel/kn.gif')

    turtle.hideturtle()
    turtle.penup()
    turtle.goto(-100, 310)
    turtle.pendown()
    turtle.color('white')
    turtle.write("Wordle", font=('arial', 70, 'bold'))
    mode = 1
    x1 = 'Normal'
    for i in range(2):
        ttl4.speed('fastest')
        ttl4.penup()
        ttl4.setposition(a2 + i * 200, 0)
        ttl4.pendown()
        ttl4.hideturtle()
        ttl4.color('white')
        ttl4.width(2)
        ttl4.forward(200)
        ttl4.left(90)
        ttl4.forward(200)
        ttl4.left(90)
        ttl4.forward(200)
        ttl4.left(90)
        ttl4.forward(200)
        ttl4.left(90)
        ttl4.penup()
        ttl4.goto(p3 + i * 200, p4)
        ttl4.write(f'{x1} \n  Mode', font=('arial', size2, 'bold'))
        ttl4.pendown()
        x1 = " Genre"


def genre():
    ttl.clear()
    ttl3.clear()

    global mode1
    mode1 = 1
    glist = ['Food', "Sports", "Places", 'Movies']
    for i in range(4):
        ttl4.penup()
        ttl4.goto(-300 + i * 150, 0)
        ttl4.pendown()
        ttl4.forward(150)
        ttl4.left(90)
        ttl4.forward(150)
        ttl4.left(90)
        ttl4.forward(150)
        ttl4.left(90)
        ttl4.forward(150)
        ttl4.left(90)
        ttl4.penup()
        ttl4.goto(-220 + i * 150, 60)
        ttl4.pendown()
        ttl4.write(glist[i], align='center', font=('arial', 30, 'bold'))

def drawno(n,random_s, list):
    global game,food,places,movies,sports
    ttl.clear()
    ttl3.clear()
    b = -700
    c = 200
    available = []

    for j in range(6):
        for i in range(n):
            ttl.speed('fastest')
            ttl.penup()
            ttl.setposition(b, c)
            ttl.pendown()
            ttl.hideturtle()
            ttl.color('white')
            ttl.width(2)
            ttl.forward(100)
            ttl.left(90)
            ttl.forward(100)
            ttl.left(90)
            ttl.forward(100)
            ttl.left(90)
            ttl.forward(100)
            ttl.left(90)
            b = b + 100
        b = -700

        c = c - 100
    c = 200
    for j in range(6):
        a = ' '
        ttl3.hideturtle()
        while (a not in english_words_set and a not in list) or len(a) != n:
            a = turtle.textinput("Enter the word", "word")
            if a is None:
                str3 = 'You canceled!!!'
                str2 = 'The word is ' + random_s
                h1 = tkinter.Tk()
                h1.title("Answer")
                defi = ''
                try:
                    response = requests.get(f'https://api.dictionaryapi.dev/api/v2/entries/en/{random_s}')
                    print(response.json()[0]['meanings'][0]['definitions'][0]['definition'])
                    defi = response.json()[0]['meanings'][0]['definitions'][0]['definition']
                except KeyError:
                    pass

                tkinter.Label(h1, text=str3, font=("Times New Roman", 30, "bold")).pack(pady=50)
                tkinter.Label(h1, text=str2, font=("Times New Roman", 18, "bold"), justify="left").pack(padx=100)
                tkinter.Label(h1, text=defi, font=("Times New Roman", 18, "bold"), justify="left").pack(pady=50)

                gamemode()
                h1.mainloop()
                game=0
                break
        for i in range(n):
            ttl3.color('black')
            ttl3.penup()
            ttl3.goto(b + 50, c + 20)
            ttl3.pendown()
            ttl3.color('white')
            ttl3.write(a[i], align="center", font=('arial', 50, 'bold'))
            b = b + 100
        b = -700
        for i in range(n):
            ttl3.speed('fastest')
            ttl3.hideturtle()
            ttl3.penup()
            ttl3.setposition(b, c)
            ttl3.pendown()
            ttl3.begin_fill()
            if a[i] == random_s[i]:
                ttl3.fillcolor('green')
            elif a[i] in random_s:
                ttl3.fillcolor('yellow')
            else:
                ttl3.fillcolor('gray')
            ttl3.width(2)
            ttl3.pendown()
            ttl3.forward(100)
            ttl3.left(90)
            ttl3.forward(100)
            ttl3.left(90)
            ttl3.forward(100)
            ttl3.left(90)
            ttl3.forward(100)
            ttl3.left(90)
            ttl3.penup()
            ttl3.end_fill()
            ttl3.penup()
            ttl3.goto(b + 50, c + 20)
            ttl3.pendown()
            ttl3.color('black')
            ttl3.write(a[i], align="center", font=('arial', 50, 'bold'))
            ttl3.color('white')
            b = b + 100
        if a.lower() == random_s:
            str3='You won!!! '
            str2='The word is '+random_s
            h1 = tkinter.Tk()
            h1.title("Answer")
            defi=''
            try:
                response = requests.get(f'https://api.dictionaryapi.dev/api/v2/entries/en/{random_s}')
                print(response.json()[0]['meanings'][0]['definitions'][0]['definition'])
                defi=response.json()[0]['meanings'][0]['definitions'][0]['definition']
            except KeyError:
                pass

            tkinter.Label(h1, text=str3, font=("Times New Roman", 30, "bold")).pack(pady=50)
            tkinter.Label(h1, text=str2, font=("Times New Roman", 18, "bold"), justify="left").pack(padx=100)
            tkinter.Label(h1, text=defi, font=("Times New Roman", 18, "bold"), justify="left").pack( pady=50)


            gamemode()
            game = 0
            h1.mainloop()
            break
        elif j == 5:
            str3='You lost!!!'
            str2='The word is '+random_s
            h1 = tkinter.Tk()
            h1.title("Answer")
            defi=''
            try:
                response = requests.get(f'https://api.dictionaryapi.dev/api/v2/entries/en/{random_s}')
                print(response.json()[0]['meanings'][0]['definitions'][0]['definition'])
                defi=response.json()[0]['meanings'][0]['definitions'][0]['definition']
            except KeyError:
                pass

            tkinter.Label(h1, text=str3, font=("Times New Roman", 30, "bold")).pack(pady=50)
            tkinter.Label(h1, text=str2, font=("Times New Roman", 18, "bold"), justify="left").pack(padx=100)
            tkinter.Label(h1, text=defi, font=("Times New Roman", 18, "bold"), justify="left").pack()


            gamemode()
            game = 0
            break
        c = c - 100
        b = -700


turtle.bgcolor("black")


def choice():
    ttl.clear()
    ttl3.clear()
    ttl1.hideturtle()
    ttl1.color("white")
    ttl1.penup()
    ttl1.goto(p1, p2)
    ttl1.pendown()
    ttl1.write("Choose the Number of Letters", align="center", font=('arial', size1, 'bold'))
    for i in range(3):
        ttl1.speed(10)
        ttl1.penup()
        ttl1.setposition(a1 + (i) * 100, b1)
        ttl1.pendown()
        ttl1.color('white')
        ttl1.width(2)
        ttl1.forward(100)
        ttl1.left(90)
        ttl1.forward(100)
        ttl1.left(90)
        ttl1.forward(100)
        ttl1.left(90)
        ttl1.forward(100)
        ttl1.left(90)
        ttl1.penup()
        ttl1.setposition(a1 + i * 100 + 50, b1 + 40)
        ttl1.pendown()
        ttl1.write(i + 4, align="center", font=("arial", 20, 'bold'))


def button_click(x, y):
    global n, a2, b2, p3, p4, size2, game, mode, mode1, food, places, movies, sports
    global  a1, b1, p1, p2, size1
    if -300 < x < -150 and 0 < y < 150 and mode1 == 1:
        game=1
        ttl4.clear()
        mode1 = 0
        food = 1
        choice()
    if -150 < x < 0 and 0 < y < 150 and mode1 == 1:
        game=1
        ttl4.clear()
        mode1 = 0
        sports = 1
        choice()
    if 0 < x < 150 and 0 < y < 150 and mode1 == 1:
        game=1
        ttl4.clear()
        mode1 = 0
        places = 1
        choice()
    if 150 < x < 300 and 0 < y < 150 and mode1 == 1:
        game=1
        ttl4.clear()
        mode1 = 0
        movies = 1
        choice()
    if a2 < x < a2+200 and 0 < y < 200 and mode == 1:
        game=1
        ttl4.clear()
        mode = 0
        choice()
    elif a2+200 < x < a2+400 and 0 < y < 200 and mode == 1:
        ttl4.clear()
        mode = 0
        genre()
    if 650 < x < 750 and 200 < y < 300:
        print("work")
        h = tkinter.Tk()
        h.title("HELP")

        tkinter.Label(h, text="How to Play:", font=("Times New Roman", 18, "bold")).pack()
        tkinter.Label(h, text=str1, font=("Times New Roman", 18, "bold"), justify="left").pack(padx=100, pady=50)
        tkinter.Label(h, text="", font=("Times New Roman", 18, "bold"), justify="left",use=True).pack(padx=100, pady=50)
        h.mainloop()
    if a1 < x < a1 + 100 and b1 < y < b1 + 100 and game == 1:
        n = 4
        game=0
        if food == 1:
            random_s = list4[random.randrange(0, len(list4))]
            l = list4
        elif places==1:
            random_s= list7[random.randrange(0, len(list7))]
            l = list7
        elif sports==1:
            random_s= list10[random.randrange(0, len(list10))]
            l = list10
        elif movies==1:
            random_s= list13[random.randrange(0, len(list13))]
            l = list13
        else:
            random_s = list2[random.randrange(0, len(list2))]
            l = list2
        print(random_s)
        sports,places,food,movies=0,0,0,0
    elif a1 + 100 < x < a1 + 200 and b1 < y < b1 + 100 and game == 1:
        n = 5
        game=0
        if food == 1:
            random_s = list5[random.randrange(0, len(list4))]
            l = list5
        elif places == 1:
            random_s = list8[random.randrange(0, len(list8))]
            l = list5
        elif sports == 1:
            random_s = list11[random.randrange(0, len(list11))]
            l = list11
        elif movies==1:
            random_s= list14[random.randrange(0, len(list14))]
            l = list14
        else:
            random_s = list1[random.randrange(0, len(list1))]
            l = list1
        sports,places,food,movies=0,0,0,0
        print(random_s)
    elif a1 + 200 < x < a1 + 300 and b1 < y < b1 + 100 and game == 1:
        n = 6
        game=0
        if food == 1:
            random_s = list6[random.randrange(0, len(list4))]
            l = list6
        elif places==1:
            random_s= list9[random.randrange(0, len(list9))]
            l = list9
        elif sports==1:
            random_s= list12[random.randrange(0, len(list12))]
            l = list12
        elif movies == 1:
            random_s = list15[random.randrange(0, len(list15))]
            l = list7
        else:
            random_s = list3[random.randrange(0, len(list3))]
            l = list3
        sports,places,food,movies=0,0,0,0
        print(random_s)
    else:
        return 0
    a2 = 100
    p3=110
    p4=30

    size2=40

    ttl1.clear()
    drawno(n, random_s, l)


def help():
    ttl2.hideturtle()
    ttl2.speed(10)
    ttl2.color("white")
    ttl2.penup()
    ttl2.goto(650, 200)
    # ttl2.pendown()
    ttl2.forward(100)
    ttl2.left(90)
    ttl2.forward(100)
    ttl2.left(90)
    ttl2.forward(100)
    ttl2.left(90)
    ttl2.forward(100)
    ttl2.back(30)
    ttl2.pendown()
    ttl2.pensize(2)
    ttl2.circle(50)
    ttl2.write("    ?   ", False, align="left", font=("Arial", 27, "bold"))
    ttl2.circle(50, 45)
    ttl2.write("HELP", False, align="left", font=("Arial", 22, "italic"))


help()
a1 = -150
b1 = -50
p1 = 0
p2 = 100
size1 = 40
gamemode()
turtle.Screen().onclick(button_click)
turtle.done()
